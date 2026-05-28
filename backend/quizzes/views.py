from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db.models import Count, Max
from django.shortcuts import get_object_or_404
from rest_framework import permissions, status, viewsets, parsers
from rest_framework.decorators import action
import csv
from io import StringIO
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse
from django.db import transaction
import openpyxl
from openpyxl.styles import Font, PatternFill

from quizzes.models import Quiz, QuizRegistration, Question, Choice, QuizAttempt, StudentAnswer, FFFAnswer, HotseatAttempt, SwitchCategory, SystemPreferences
from quizzes.serializers import QuizRegistrationSerializer, QuizSerializer, FFFAnswerSerializer, HotseatAttemptSerializer, QuestionSerializer, EnrolledStudentSerializer, SystemPreferencesSerializer
from quizzes.services import process_mock_payment, register_student_for_quiz
from django.utils import timezone
import random

User = get_user_model()


class IsAdminUser(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and request.user.role == 'admin')


class IsStudentUser(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and request.user.role == 'student')


class AdminQuizViewSet(viewsets.ModelViewSet):
    """
    CRUD for admin to manage quizzes.
    Includes archived quizzes by default.
    """
    permission_classes = [IsAdminUser]
    serializer_class = QuizSerializer
    
    def get_queryset(self):
        return Quiz.objects.annotate(
            registered_count=Count('registrations')
        ).all()
        
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        template_type = request.query_params.get('type', 'prelim').strip().lower()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if template_type == 'fff':
            ws.title = "FFF Sequencing Template"
            headers = [
                'Question Text', 'Option A', 'Option B', 'Option C', 'Option D', 
                'Correct Sequence (e.g. CADB)', 'Marks', 
                'Question Type (fff_1/fff_2/fff_3)', 'Category', 'Trivia'
            ]
            sample_row = [
                "Arrange these Indian monuments in chronological order of construction (earliest first):",
                "Taj Mahal",
                "Red Fort",
                "Qutub Minar",
                "Gateway of India",
                "CADB",
                1,
                "fff_1",
                "History",
                "Qutub Minar (1199) -> Taj Mahal (1632) -> Red Fort (1638) -> Gateway of India (1911)."
            ]
            filename = "fff_sequencing_template.xlsx"
        elif template_type == 'hotseat':
            ws.title = "Hotseat MCQ Template"
            headers = [
                'Question Text', 'Option A', 'Option B', 'Option C', 'Option D', 
                'Correct Option (A/B/C/D)', 'Marks', 
                'Question Type (hotseat_1/hotseat_2/hotseat_3)', 'Category', 'Trivia'
            ]
            sample_row = [
                "What is the chemical formula of Table Salt?",
                "HCl",
                "H2O",
                "NaCl",
                "CO2",
                "C",
                1,
                "hotseat_1",
                "Science",
                "NaCl stands for Sodium Chloride which is common table salt."
            ]
            filename = "hotseat_quiz_template.xlsx"
        else:
            ws.title = "Preliminary MCQ Template"
            headers = [
                'Question Text', 'Option A', 'Option B', 'Option C', 'Option D', 
                'Correct Option (A/B/C/D)', 'Marks', 
                'Question Type (regular)', 'Category', 'Trivia'
            ]
            sample_row = [
                "What is the capital of France?",
                "London",
                "Berlin",
                "Paris",
                "Madrid",
                "C",
                1,
                "regular",
                "General",
                "Paris is the most populous city of France and has been one of Europe's major centres of finance, diplomacy, commerce, fashion, science and arts since the 17th century."
            ]
            filename = "preliminary_quiz_template.xlsx"
            
        ws.append(headers)
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        ws.append(sample_row)
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @action(detail=True, methods=['post'], parser_classes=[parsers.MultiPartParser])
    def upload_questions(self, request, pk=None):
        quiz = self.get_object()
        if 'file' not in request.FILES:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
            
        file = request.FILES['file']
        if not file.name.endswith('.xlsx'):
            return Response({"detail": "Please upload an Excel (.xlsx) file."}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) <= 1:
                return Response({"detail": "No questions found in Excel file."}, status=status.HTTP_400_BAD_REQUEST)
                
            created_count = 0
            error_count = 0
            error_log = []
            seen_questions = set()
            
            max_order = quiz.questions.aggregate(Max('order'))['order__max'] or 0
            
            for idx, row in enumerate(rows[1:], start=2):
                if not row or not any(row):
                    continue
                    
                text = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
                if not text:
                    error_log.append({"row": idx, "question": f"Row {idx}", "reason": "Question text is missing."})
                    error_count += 1
                    continue
                    
                if quiz.questions.filter(text=text).exists():
                    error_log.append({"row": idx, "question": text[:40], "reason": "Question text already exists in this quiz."})
                    error_count += 1
                    continue
                    
                if text in seen_questions:
                    error_log.append({"row": idx, "question": text[:40], "reason": "Question text is duplicated inside the Excel file."})
                    error_count += 1
                    continue
                seen_questions.add(text)
                # Detect Excel template format dynamically (Old 10-column vs New 21-column)
                is_old_format = True
                
                # Check if row[18] is a valid question type
                if len(row) > 18 and str(row[18]).strip().lower() in Question.QuestionType.values:
                    is_old_format = False
                elif len(row) > 7 and str(row[7]).strip().lower() in Question.QuestionType.values:
                    is_old_format = True
                else:
                    # Fallback default: if the row has more than 12 columns, assume new format
                    is_old_format = (len(row) < 12)
                
                if is_old_format:
                    # Old 10-column format
                    opt_a = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                    opt_b = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                    opt_c = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
                    opt_d = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
                    
                    raw_options = [opt_a, opt_b, opt_c, opt_d] + [''] * 11
                    
                    correct_opt = str(row[5]).strip() if len(row) > 5 and row[5] is not None else 'A'
                    marks = int(row[6]) if len(row) > 6 and row[6] is not None else 1
                    q_type = str(row[7]).strip().lower() if len(row) > 7 and row[7] else 'regular'
                    category = str(row[8]).strip() if len(row) > 8 and row[8] else 'General'
                    trivia = str(row[9]).strip() if len(row) > 9 and row[9] is not None else ''
                else:
                    # New 21-column format
                    raw_options = []
                    for o_idx in range(15):
                        col_val = str(row[1 + o_idx]).strip() if len(row) > (1 + o_idx) and row[1 + o_idx] is not None else ''
                        raw_options.append(col_val)
                        
                    correct_opt = str(row[16]).strip() if len(row) > 16 and row[16] is not None else 'A'
                    marks = int(row[17]) if len(row) > 17 and row[17] is not None else 1
                    q_type = str(row[18]).strip().lower() if len(row) > 18 and row[18] else 'regular'
                    category = str(row[19]).strip() if len(row) > 19 and row[19] else 'General'
                    trivia = str(row[20]).strip() if len(row) > 20 and row[20] is not None else ''
                
                # Filter out trailing empty options to find the non-empty option set
                non_empty_options = [o for o in raw_options if o]
                num_options = len(non_empty_options)
                
                # Ensure options are contiguous without gaps (only check up to num_options)
                has_gap = False
                for idx_opt in range(num_options):
                    if not raw_options[idx_opt]:
                        has_gap = True
                        break
                if has_gap:
                    error_log.append({"row": idx, "question": text[:40], "reason": "Options must be contiguous starting from Option A without empty slots."})
                    error_count += 1
                    continue
                
                if q_type not in Question.QuestionType.values:
                    q_type = 'regular'
                
                # Validation checks specific to FFF vs regular MCQ
                if q_type.startswith('fff_'):
                    if num_options < 8 or num_options > 15:
                        error_log.append({"row": idx, "question": text[:40], "reason": f"Fastest Finger First questions must have between 8 and 15 options. Got {num_options}."})
                        error_count += 1
                        continue
                    
                    clean_seq = [ch.upper() for ch in correct_opt if ch.isalpha()]
                    expected_letters = [chr(ord('A') + i) for i in range(num_options)]
                    
                    if len(clean_seq) != num_options or set(clean_seq) != set(expected_letters):
                        error_log.append({"row": idx, "question": text[:40], "reason": f"Correct sequence for FFF must specify exactly the non-empty option letters A to {chr(ord('A') + num_options - 1)} in order (case-insensitive). Got '{correct_opt}'."})
                        error_count += 1
                        continue
                    
                    seq_map = {letter: rank for rank, letter in enumerate(clean_seq, 1)}
                else:
                    # MCQ strictly requires exactly 4 options
                    if num_options != 4:
                        error_log.append({"row": idx, "question": text[:40], "reason": f"Standard MCQ / Preliminary / Hotseat questions must have exactly 4 options. Got {num_options}."})
                        error_count += 1
                        continue
                    
                    correct_opt_upper = correct_opt.upper()
                    expected_letters = ['A', 'B', 'C', 'D']
                    if correct_opt_upper not in expected_letters:
                        error_log.append({"row": idx, "question": text[:40], "reason": f"Correct option letter '{correct_opt}' is invalid. Must be A, B, C, or D (case-insensitive)."})
                        error_count += 1
                        continue
                
                with transaction.atomic():
                    question = Question.objects.create(
                        quiz=quiz, text=text, order=max_order + created_count + 1, marks=marks,
                        question_type=q_type, category=category, trivia=trivia
                    )
                    
                    if q_type.startswith('fff_'):
                        for letter_idx, letter in enumerate([chr(ord('A') + i) for i in range(num_options)]):
                            Choice.objects.create(
                                question=question,
                                text=raw_options[letter_idx],
                                is_correct=False,
                                correct_order=seq_map[letter]
                            )
                    else:
                        for letter_idx, letter in enumerate(expected_letters):
                            Choice.objects.create(
                                question=question,
                                text=raw_options[letter_idx],
                                is_correct=(correct_opt_upper == letter),
                                correct_order=None
                            )
                    
                created_count += 1
                
            return Response({
                "detail": f"Successfully imported {created_count} questions.",
                "success_count": created_count,
                "error_count": error_count,
                "errors": error_log
            })
            
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get'])
    def questions(self, request, pk=None):
        quiz = self.get_object()
        questions = quiz.questions.all().prefetch_related('choices')
        
        data = []
        for q in questions:
            choices_data = []
            for c in q.choices.all():
                choices_data.append({
                    "id": c.id,
                    "text": c.text,
                    "is_correct": c.is_correct,
                    "correct_order": c.correct_order
                })
            data.append({
                "id": q.id,
                "text": q.text,
                "order": q.order,
                "marks": q.marks,
                "question_type": q.question_type,
                "category": q.category,
                "trivia": q.trivia,
                "choices": choices_data
            })
        return Response(data)

    @action(detail=True, methods=['get'])
    def switch_categories(self, request, pk=None):
        quiz = self.get_object()
        categories = quiz.switch_categories.all().select_related('question')
        
        data = []
        for c in categories:
            choices_data = []
            q_data = None
            if c.question:
                for choice in c.question.choices.all():
                    choices_data.append({
                        "id": choice.id,
                        "text": choice.text,
                        "is_correct": choice.is_correct
                    })
                q_data = {
                    "id": c.question.id,
                    "text": c.question.text,
                    "choices": choices_data
                }
                
            img_url = c.image.url if c.image else None
            if img_url and request:
                img_url = request.build_absolute_uri(img_url)
                
            data.append({
                "id": c.id,
                "name": c.name,
                "image": img_url,
                "question": q_data
            })
        return Response(data)

    @action(detail=True, methods=['post'], parser_classes=[parsers.MultiPartParser, parsers.FormParser])
    def save_switch_category(self, request, pk=None):
        quiz = self.get_object()
        
        category_id = request.data.get('category_id')
        if not category_id and quiz.switch_categories.count() >= 6:
            return Response({"detail": "You can configure a maximum of 6 switch categories."}, status=400)
            
        name = request.data.get('name', '').strip()
        if not name:
            return Response({"detail": "Category name is required."}, status=400)
            
        question_text = request.data.get('question_text', '').strip()
        if not question_text:
            return Response({"detail": "Question text is required."}, status=400)
            
        choice_a = request.data.get('choice_a', '').strip()
        choice_b = request.data.get('choice_b', '').strip()
        choice_c = request.data.get('choice_c', '').strip()
        choice_d = request.data.get('choice_d', '').strip()
        correct_choice = request.data.get('correct_choice', '').strip().upper()
        
        if not all([choice_a, choice_b, choice_c, choice_d]) or correct_choice not in ['A', 'B', 'C', 'D']:
            return Response({"detail": "All 4 options and a correct selection (A/B/C/D) are required."}, status=400)
            
        with transaction.atomic():
            if category_id:
                category = get_object_or_404(SwitchCategory, quiz=quiz, id=category_id)
                category.name = name
                if 'image' in request.FILES:
                    category.image = request.FILES['image']
                category.save()
            else:
                image_file = request.FILES.get('image')
                category = SwitchCategory.objects.create(
                    quiz=quiz,
                    name=name,
                    image=image_file
                )
                
            if category.question:
                question = category.question
                question.text = question_text
                question.category = name
                question.save()
                question.choices.all().delete()
            else:
                question = Question.objects.create(
                    quiz=quiz,
                    text=question_text,
                    question_type=Question.QuestionType.SWITCH,
                    category=name,
                    order=0
                )
                category.question = question
                category.save()
                
            Choice.objects.create(question=question, text=choice_a, is_correct=(correct_choice == 'A'))
            Choice.objects.create(question=question, text=choice_b, is_correct=(correct_choice == 'B'))
            Choice.objects.create(question=question, text=choice_c, is_correct=(correct_choice == 'C'))
            Choice.objects.create(question=question, text=choice_d, is_correct=(correct_choice == 'D'))
            
        return Response({"detail": "Switch category and question saved successfully."})

    @action(detail=True, methods=['post'])
    def delete_switch_category(self, request, pk=None):
        quiz = self.get_object()
        category_id = request.data.get('category_id')
        if not category_id:
            return Response({"detail": "Category ID is required."}, status=400)
            
        category = get_object_or_404(SwitchCategory, quiz=quiz, id=category_id)
        with transaction.atomic():
            if category.question:
                category.question.delete()
            category.delete()
            
        return Response({"detail": "Switch category deleted successfully."})

    @action(detail=True, methods=['post'])
    def add_question(self, request, pk=None):
        quiz = self.get_object()
        text = request.data.get('text', '').strip()
        if not text:
            return Response({"detail": "Question text is required."}, status=400)
            
        q_type = request.data.get('question_type', 'regular')
        category = request.data.get('category', 'General')
        marks = int(request.data.get('marks', 1))
        trivia = request.data.get('trivia', '')
        choices_data = request.data.get('choices', [])
        
        max_order = quiz.questions.aggregate(Max('order'))['order__max'] or 0
        
        with transaction.atomic():
            question = Question.objects.create(
                quiz=quiz, text=text, order=max_order + 1,
                marks=marks, question_type=q_type, category=category, trivia=trivia
            )
            for c in choices_data:
                Choice.objects.create(
                    question=question,
                    text=c.get('text', '').strip(),
                    is_correct=c.get('is_correct', False),
                    correct_order=c.get('correct_order')
                )
        return Response({"detail": "Question added successfully.", "id": question.id})

    @action(detail=False, methods=['post'])
    def edit_question(self, request):
        question_id = request.data.get('id')
        question = get_object_or_404(Question, id=question_id)
        
        text = request.data.get('text', '').strip()
        if not text:
            return Response({"detail": "Question text is required."}, status=400)
            
        q_type = request.data.get('question_type', 'regular')
        category = request.data.get('category', 'General')
        marks = int(request.data.get('marks', 1))
        trivia = request.data.get('trivia', '')
        choices_data = request.data.get('choices', [])
        
        with transaction.atomic():
            question.text = text
            question.question_type = q_type
            question.category = category
            question.marks = marks
            question.trivia = trivia
            question.save()
            
            question.choices.all().delete()
            for c in choices_data:
                Choice.objects.create(
                    question=question,
                    text=c.get('text', '').strip(),
                    is_correct=c.get('is_correct', False),
                    correct_order=c.get('correct_order')
                )
        return Response({"detail": "Question updated successfully."})

    @action(detail=False, methods=['post'])
    def delete_question(self, request):
        question_id = request.data.get('id')
        question = get_object_or_404(Question, id=question_id)
        question.delete()
        return Response({"detail": "Question deleted successfully."})

    @action(detail=True, methods=['post'])
    def update_stage(self, request, pk=None):
        quiz = self.get_object()
        stage = request.data.get('stage')
        if not stage or stage not in Quiz.Stage.values:
            return Response({"detail": "Invalid stage provided."}, status=400)
        quiz.current_stage = stage
        quiz.save(update_fields=['current_stage'])
        return Response(QuizSerializer(quiz).data)

    @action(detail=True, methods=['get'])
    def host_hotseat_question(self, request, pk=None):
        """Admin host view: see current question with correct answers, trivia, and contestant's preselection."""
        quiz = self.get_object()
        stage = quiz.current_stage
        
        hotseat_player = None
        batch_num = None
        q_type = None
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
            q_type = Question.QuestionType.HOTSEAT_1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
            q_type = Question.QuestionType.HOTSEAT_2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
            q_type = Question.QuestionType.HOTSEAT_3
        else:
            return Response({"active": False, "detail": "No hotseat stage active."})
        
        if not hotseat_player:
            return Response({"active": False, "detail": "No hotseat player promoted."})
        
        attempt = HotseatAttempt.objects.filter(quiz=quiz, student=hotseat_player, batch_number=batch_num).first()
        if not attempt:
            return Response({"active": False, "detail": "No hotseat attempt found."})
        
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({
                "active": False, "completed": True,
                "status": attempt.status, "score": attempt.score,
                "contestant_name": hotseat_player.full_name
            })
        
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        if attempt.current_question_index >= len(questions):
            return Response({"active": False, "completed": True, "status": "completed", "score": attempt.score})
        
        question = questions[attempt.current_question_index]
        choices = list(question.choices.all())
        
        return Response({
            "active": True,
            "current_index": attempt.current_question_index,
            "total_questions": len(questions),
            "score": attempt.score,
            "contestant_name": hotseat_player.full_name,
            "preselected_choice_id": attempt.preselected_choice_id,
            "lifelines": {
                "5050_used": attempt.lifeline_5050_used,
                "poll_used": attempt.lifeline_poll_used,
                "switch_used": attempt.lifeline_switch_used
            },
            "lifeline_request_status": attempt.lifeline_request_status,
            "pending_lifeline_type": attempt.pending_lifeline_type,
            "pending_lifeline_switch_category": attempt.pending_lifeline_switch_category,
            "approved_lifeline_data": attempt.approved_lifeline_data,
            "timer_is_paused": attempt.timer_is_paused,
            "options_visible": attempt.options_visible,
            "showing_question": attempt.showing_question,
            "question": {
                "id": question.id,
                "text": question.text,
                "category": question.category,
                "trivia": question.trivia,
                "choices": [{"id": c.id, "text": c.text, "is_correct": c.is_correct} for c in choices]
            }
        })

    @action(detail=True, methods=['post'])
    def host_lock_answer(self, request, pk=None):
        """Admin host action: lock the contestant's preselected answer and process scoring."""
        quiz = self.get_object()
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
            q_type = Question.QuestionType.HOTSEAT_1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
            q_type = Question.QuestionType.HOTSEAT_2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
            q_type = Question.QuestionType.HOTSEAT_3
        else:
            return Response({"detail": "No hotseat stage active."}, status=400)
        
        if not hotseat_player:
            return Response({"detail": "No hotseat player promoted."}, status=400)
        
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=hotseat_player, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
        
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        if attempt.current_question_index >= len(questions):
            return Response({"detail": "All questions completed."}, status=400)
        
        question = questions[attempt.current_question_index]
        choice_id = attempt.preselected_choice_id
        if not choice_id:
            return Response({"detail": "Contestant has not selected any option yet."}, status=400)
        
        selected_choice = Choice.objects.filter(question=question, id=choice_id).first()
        is_correct = selected_choice.is_correct if selected_choice else False
        correct_choice = Choice.objects.filter(question=question, is_correct=True).first()
        
        if is_correct:
            current_points = SCORE_LADDER[attempt.current_question_index] if attempt.current_question_index < len(SCORE_LADDER) else 0
            attempt.score = current_points
            attempt.current_question_index += 1
            attempt.preselected_choice = None
            attempt.current_question_switched = False
            
            # Reset option visibility and hold next question until explicitly pushed by Host
            attempt.showing_question = False
            attempt.options_visible = False
            attempt.timer_is_paused = False
            
            completed = attempt.current_question_index >= len(questions)
            if completed:
                attempt.status = HotseatAttempt.Status.COMPLETED
                attempt.completed_at = timezone.now()
                self._host_save_score(quiz, batch_num, attempt.score, "completed")
            else:
                self._host_save_score(quiz, batch_num, attempt.score, "playing")
            
            attempt.save()
            return Response({
                "correct": True,
                "correct_choice_id": correct_choice.id if correct_choice else None,
                "selected_choice_id": choice_id,
                "current_points": attempt.score,
                "next_index": attempt.current_question_index,
                "completed": completed,
                "trivia": question.trivia,
                "message": f"All {len(questions)} questions completed! Final score: {attempt.score} pts" if completed else f"Correct! {hotseat_player.full_name} has reached Question {attempt.current_question_index + 1}!"
            })
        else:
            checkpoint_score = 0
            fail_index = attempt.current_question_index
            if fail_index >= 10:
                checkpoint_score = 100
            elif fail_index >= 5:
                checkpoint_score = 50
            
            attempt.score = checkpoint_score
            attempt.status = HotseatAttempt.Status.FAILED
            attempt.completed_at = timezone.now()
            attempt.preselected_choice = None
            attempt.save()
            
            self._host_save_score(quiz, batch_num, attempt.score, "failed")
            
            return Response({
                "correct": False,
                "correct_choice_id": correct_choice.id if correct_choice else None,
                "selected_choice_id": choice_id,
                "checkpoint_points": checkpoint_score,
                "completed": True,
                "trivia": question.trivia,
                "message": f"Incorrect! The correct answer was '{correct_choice.text if correct_choice else 'N/A'}'. Score drops to checkpoint: {checkpoint_score} pts"
            })

    def _host_save_score(self, quiz, batch_num, score, status_str):
        if batch_num == 1:
            quiz.hotseat_score_1 = score
            quiz.hotseat_status_1 = status_str
            quiz.save(update_fields=['hotseat_score_1', 'hotseat_status_1'])
        elif batch_num == 2:
            quiz.hotseat_score_2 = score
            quiz.hotseat_status_2 = status_str
            quiz.save(update_fields=['hotseat_score_2', 'hotseat_status_2'])
        elif batch_num == 3:
            quiz.hotseat_score_3 = score
            quiz.hotseat_status_3 = status_str
            quiz.save(update_fields=['hotseat_score_3', 'hotseat_status_3'])

    @action(detail=True, methods=['post'])
    def set_batches(self, request, pk=None):
        quiz = self.get_object()
        batch_1 = request.data.get('batch_1', [])
        batch_2 = request.data.get('batch_2', [])
        batch_3 = request.data.get('batch_3', [])
        
        if not batch_1 or not batch_2 or not batch_3:
            attempts = QuizAttempt.objects.filter(quiz=quiz, completed_at__isnull=False).order_by('-score', 'completed_at')
            student_ids = [att.student_id for att in attempts]
            
            top_30 = student_ids[:30]
            
            batch_1 = top_30[0:10]
            batch_2 = top_30[10:20]
            batch_3 = top_30[20:30]
            
            quiz.top_30_selected = top_30
        else:
            quiz.top_30_selected = list(batch_1) + list(batch_2) + list(batch_3)

        quiz.batch_1_players = batch_1
        quiz.batch_2_players = batch_2
        quiz.batch_3_players = batch_3
        quiz.save(update_fields=['top_30_selected', 'batch_1_players', 'batch_2_players', 'batch_3_players'])
        return Response(QuizSerializer(quiz).data)

    @action(detail=True, methods=['get'])
    def fff_results(self, request, pk=None):
        quiz = self.get_object()
        stage = quiz.current_stage
        if stage == Quiz.Stage.FFF_BATCH_1 or stage == Quiz.Stage.HOTSEAT_BATCH_1:
            batch_num = 1
            q_type = Question.QuestionType.FFF_1
        elif stage == Quiz.Stage.FFF_BATCH_2 or stage == Quiz.Stage.HOTSEAT_BATCH_2:
            batch_num = 2
            q_type = Question.QuestionType.FFF_2
        elif stage == Quiz.Stage.FFF_BATCH_3 or stage == Quiz.Stage.HOTSEAT_BATCH_3:
            batch_num = 3
            q_type = Question.QuestionType.FFF_3
        else:
            return Response({"detail": "FFF is not active or completed for this quiz stage."}, status=400)
            
        fff_question = Question.objects.filter(quiz=quiz, question_type=q_type).first()
        if not fff_question:
            return Response({"detail": "FFF question not found for this batch."}, status=404)
            
        answers = FFFAnswer.objects.filter(question=fff_question, batch_number=batch_num).select_related('student')
        
        serialized = FFFAnswerSerializer(answers, many=True).data
        
        for ans_data, ans_obj in zip(serialized, answers):
            ans_data['is_correct'] = ans_obj.is_correct or (ans_obj.selected_choice.is_correct if ans_obj.selected_choice else False)
            
        serialized.sort(key=lambda x: (not x['is_correct'], x['time_taken_seconds']))
        
        return Response({
            "question": QuestionSerializer(fff_question).data,
            "results": serialized
        })

    @action(detail=True, methods=['post'])
    def promote_hotseat(self, request, pk=None):
        quiz = self.get_object()
        stage = quiz.current_stage
        student_id = request.data.get('student_id')
        if not student_id:
            return Response({"detail": "Student ID is required."}, status=400)
            
        student = get_object_or_404(User, id=student_id)
        
        if stage == Quiz.Stage.FFF_BATCH_1 or stage == Quiz.Stage.HOTSEAT_BATCH_1:
            quiz.hotseat_player_1 = student
            quiz.hotseat_status_1 = "playing"
            quiz.save(update_fields=['hotseat_player_1', 'hotseat_status_1'])
            HotseatAttempt.objects.get_or_create(quiz=quiz, student=student, batch_number=1)
        elif stage == Quiz.Stage.FFF_BATCH_2 or stage == Quiz.Stage.HOTSEAT_BATCH_2:
            quiz.hotseat_player_2 = student
            quiz.hotseat_status_2 = "playing"
            quiz.save(update_fields=['hotseat_player_2', 'hotseat_status_2'])
            HotseatAttempt.objects.get_or_create(quiz=quiz, student=student, batch_number=2)
        elif stage == Quiz.Stage.FFF_BATCH_3 or stage == Quiz.Stage.HOTSEAT_BATCH_3:
            quiz.hotseat_player_3 = student
            quiz.hotseat_status_3 = "playing"
            quiz.save(update_fields=['hotseat_player_3', 'hotseat_status_3'])
            HotseatAttempt.objects.get_or_create(quiz=quiz, student=student, batch_number=3)
        else:
            return Response({"detail": "Promotion is not allowed in this stage."}, status=400)
            
        return Response(QuizSerializer(quiz).data)

    @action(detail=True, methods=['get'])
    def prelim_scores(self, request, pk=None):
        quiz = self.get_object()
        # Fetch all attempts (both completed and in-progress), ordered by:
        # 1. Completed first, then in-progress
        # 2. Higher score first
        # 3. Earlier completion time first (for completed attempts)
        from django.db.models import Case, When, Value, IntegerField
        attempts = QuizAttempt.objects.filter(quiz=quiz).select_related('student').annotate(
            completion_order=Case(
                When(completed_at__isnull=False, then=Value(0)),
                default=Value(1),
                output_field=IntegerField()
            )
        ).order_by('completion_order', '-score', 'completed_at')[:30]
        
        total_regular_questions = quiz.questions.filter(question_type=Question.QuestionType.REGULAR).count()
        
        data = []
        for idx, att in enumerate(attempts, 1):
            reg = QuizRegistration.objects.filter(quiz=quiz, student=att.student).first()
            answers = att.answers.select_related('selected_choice').all()
            correct_count = sum(1 for ans in answers if ans.selected_choice and ans.selected_choice.is_correct)
            incorrect_count = sum(1 for ans in answers if not ans.selected_choice or not ans.selected_choice.is_correct)
            is_completed = att.completed_at is not None
            data.append({
                "rank": idx,
                "student_id": att.student.id,
                "student_name": att.student.full_name,
                "player_id": reg.player_id if reg else "",
                "score": att.score,
                "time_taken": (att.completed_at - att.started_at).total_seconds() if att.completed_at and att.started_at else None,
                "correct_count": correct_count,
                "incorrect_count": incorrect_count,
                "completed": is_completed,
                "questions_answered": att.current_question_index,
                "total_questions": total_regular_questions
            })
        return Response(data)

    @action(detail=True, methods=['get'])
    def enrolled_students(self, request, pk=None):
        quiz = self.get_object()
        registrations = quiz.registrations.select_related('student').all()
        serializer = EnrolledStudentSerializer(registrations, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def remove_registration(self, request, pk=None):
        quiz = self.get_object()
        registration_id = request.data.get('registration_id')
        if not registration_id:
            return Response({"detail": "registration_id is required."}, status=status.HTTP_400_BAD_REQUEST)
        
        reg = QuizRegistration.objects.filter(id=registration_id, quiz=quiz).first()
        if not reg:
            return Response({"detail": "Registration not found."}, status=status.HTTP_404_NOT_FOUND)
        
        student = reg.student
        
        # Delete any associated quiz attempt and answers so the student can start fresh
        QuizAttempt.objects.filter(quiz=quiz, student=student).delete()
        
        # Delete FFF answers if any
        FFFAnswer.objects.filter(quiz=quiz, student=student).delete()
        
        # Delete Hotseat attempts if any
        HotseatAttempt.objects.filter(quiz=quiz, student=student).delete()
        
        # Clear hotseat player references on the quiz if this student was assigned
        update_fields = []
        if quiz.hotseat_player_1 == student:
            quiz.hotseat_player_1 = None
            quiz.hotseat_status_1 = ""
            update_fields.extend(['hotseat_player_1', 'hotseat_status_1'])
        if quiz.hotseat_player_2 == student:
            quiz.hotseat_player_2 = None
            quiz.hotseat_status_2 = ""
            update_fields.extend(['hotseat_player_2', 'hotseat_status_2'])
        if quiz.hotseat_player_3 == student:
            quiz.hotseat_player_3 = None
            quiz.hotseat_status_3 = ""
            update_fields.extend(['hotseat_player_3', 'hotseat_status_3'])
        
        # Remove student from batch player lists
        if student.id in quiz.batch_1_players:
            quiz.batch_1_players.remove(student.id)
            if 'batch_1_players' not in update_fields:
                update_fields.append('batch_1_players')
        if student.id in quiz.batch_2_players:
            quiz.batch_2_players.remove(student.id)
            if 'batch_2_players' not in update_fields:
                update_fields.append('batch_2_players')
        if student.id in quiz.batch_3_players:
            quiz.batch_3_players.remove(student.id)
            if 'batch_3_players' not in update_fields:
                update_fields.append('batch_3_players')
        
        if update_fields:
            quiz.save(update_fields=update_fields)
        
        # Delete the registration itself
        reg.delete()
        
        return Response({
            "detail": f"Registration for {student.full_name} has been fully removed. All quiz data (attempts, answers, hotseat, FFF) has been cleared. They can now re-register.",
            "removed_student_name": student.full_name,
            "removed_student_email": student.email
        })

    @action(detail=True, methods=['post'])
    def enroll_student_manual(self, request, pk=None):
        quiz = self.get_object()
        email = request.data.get('email', '').strip().lower()
        full_name = request.data.get('full_name', '').strip()
        college_id = request.data.get('roll_number', '').strip() or request.data.get('college_id', '').strip()
        payment_status = request.data.get('payment_status', 'paid').strip().lower()

        if not email or not full_name or not college_id:
            return Response({"detail": "Email, full name, and roll number are required."}, status=status.HTTP_400_BAD_REQUEST)

        prefs = SystemPreferences.get_solo()
        if prefs.auto_approve_registrations:
            payment_status = QuizRegistration.PaymentStatus.PAID
        elif payment_status not in [QuizRegistration.PaymentStatus.PAID, QuizRegistration.PaymentStatus.PENDING]:
            payment_status = QuizRegistration.PaymentStatus.PAID

        college_id = college_id.upper().strip()
        email = email.strip().lower()

        # Query existing student accounts strictly by roll_number or email case-insensitively
        from django.db.models import Q
        user = User.objects.filter(
            Q(roll_number__iexact=college_id) | 
            Q(college_id__iexact=college_id) | 
            Q(email__iexact=email)
        ).first()

        if not user:
            return Response(
                {"detail": "Student account not found. Please create the student account under 'Student Accounts' first."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if user.role != User.Role.STUDENT:
            return Response(
                {"detail": "The specified user is not a student account."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            with transaction.atomic():
                # Make sure the user's roll_number and college_id are capitalized to prevent cased duplicate lookups
                changed = False
                if user.roll_number != user.roll_number.upper():
                    user.roll_number = user.roll_number.upper()
                    changed = True
                if user.college_id != user.college_id.upper():
                    user.college_id = user.college_id.upper()
                    changed = True
                if changed:
                    user.save(update_fields=['roll_number', 'college_id', 'updated_at'])

                if QuizRegistration.objects.filter(student=user, quiz=quiz).exists():
                    reg = QuizRegistration.objects.get(student=user, quiz=quiz)
                    reg.payment_status = payment_status
                    reg.save()
                    return Response({
                        "detail": "Student is already registered for this quiz. Registration status updated.",
                        "registration": EnrolledStudentSerializer(reg).data
                    })

                max_seq = QuizRegistration.objects.filter(
                    quiz=quiz,
                    sequence_number__isnull=False
                ).aggregate(Max('sequence_number'))['sequence_number__max']
                next_seq = 1 if max_seq is None else max_seq + 1

                reg = QuizRegistration.objects.create(
                    student=user,
                    quiz=quiz,
                    payment_status=payment_status,
                    sequence_number=next_seq,
                    player_id=f"PLAYER {next_seq:03d}"
                )

            return Response({
                "detail": f"Successfully enrolled student {full_name}.",
                "registration": EnrolledStudentSerializer(reg).data
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], parser_classes=[parsers.MultiPartParser])
    def bulk_enroll_students(self, request, pk=None):
        quiz = self.get_object()
        if 'file' not in request.FILES:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)
            
        file = request.FILES['file']
        filename = file.name.lower()
        
        is_xlsx = filename.endswith('.xlsx')
        is_csv = filename.endswith('.csv')
        
        if not is_xlsx and not is_csv:
            return Response({"detail": "Please upload an Excel (.xlsx) or CSV (.csv) file."}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            rows = []
            if is_xlsx:
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            elif is_csv:
                file_data = file.read().decode('utf-8')
                csv_data = csv.reader(StringIO(file_data))
                rows = list(csv_data)
 
            if len(rows) <= 1:
                return Response({"detail": "No student records found in the uploaded file."}, status=status.HTTP_400_BAD_REQUEST)
                
            enrolled_count = 0
            skipped_count = 0
            
            from users.models import School, Program, Branch, StudentProfile
            prefs = SystemPreferences.get_solo()
            
            with transaction.atomic():
                for idx, row in enumerate(rows[1:], start=1):
                    if not row or not any(row):
                        continue
                        
                    full_name = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
                    email = str(row[1]).strip().lower() if len(row) > 1 and row[1] is not None else ''
                    college_id = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                    pay_status = str(row[3]).strip().lower() if len(row) > 3 and row[3] is not None else 'paid'
                    
                    if not email or not full_name or not college_id:
                        skipped_count += 1
                        continue
                        
                    if prefs.auto_approve_registrations:
                        pay_status = QuizRegistration.PaymentStatus.PAID
                    elif pay_status not in [QuizRegistration.PaymentStatus.PAID, QuizRegistration.PaymentStatus.PENDING]:
                        pay_status = QuizRegistration.PaymentStatus.PAID
 
                    user, created = User.objects.get_or_create(
                        email=email,
                        defaults={
                            'full_name': full_name,
                            'college_id': college_id,
                            'roll_number': college_id,
                            'role': User.Role.STUDENT
                        }
                    )
                    if created:
                        user.set_password("KBC123")
                        user.save()
                    else:
                        changed = False
                        if not user.college_id:
                            user.college_id = college_id
                            changed = True
                        if not user.roll_number:
                            user.roll_number = college_id
                            changed = True
                        if changed:
                            user.save()
                            
                    if not hasattr(user, 'student_profile'):
                        school = School.objects.first()
                        if not school:
                            school = School.objects.create(school_name="Default School", school_code="DEFAULT_SCH")
                        
                        program = Program.objects.filter(school=school).first()
                        if not program:
                            program = Program.objects.create(school=school, program_name="Default Program", program_code="DEFAULT_PROG")
                        
                        branch = Branch.objects.filter(program=program).first()
                        if not branch:
                            branch = Branch.objects.create(program=program, branch_name="Default Branch", branch_code="DEFAULT_BR")
 
                        StudentProfile.objects.create(
                            user=user,
                            school=school,
                            program=program,
                            branch=branch,
                            year=StudentProfile.Year.FIRST
                        )
                        
                    if QuizRegistration.objects.filter(student=user, quiz=quiz).exists():
                        reg = QuizRegistration.objects.get(student=user, quiz=quiz)
                        reg.payment_status = pay_status
                        reg.save()
                        continue
                        
                    max_seq = QuizRegistration.objects.filter(
                        quiz=quiz,
                        sequence_number__isnull=False
                    ).aggregate(Max('sequence_number'))['sequence_number__max']
                    next_seq = 1 if max_seq is None else max_seq + 1
 
                    QuizRegistration.objects.create(
                        student=user,
                        quiz=quiz,
                        payment_status=pay_status,
                        sequence_number=next_seq,
                        player_id=f"PLAYER {next_seq:03d}"
                    )
                    enrolled_count += 1
                    
            return Response({
                "detail": f"Successfully enrolled {enrolled_count} students. Skipped {skipped_count} invalid records."
            })
            
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
 
    @action(detail=False, methods=['get'])
    def download_enrollment_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Enrollment Template"
        
        headers = ['Full Name', 'Email', 'Roll Number', 'Payment Status (paid/pending)']
        ws.append(headers)
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        sample_row = [
            "John Doe",
            "johndoe@quizverse.edu",
            "ROLL001",
            "paid"
        ]
        ws.append(sample_row)
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 40)
 
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="student_enrollment_template.xlsx"'
        wb.save(response)
        return response


class MyQuizRegistrationView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        reg = QuizRegistration.objects.filter(quiz=quiz, student=request.user).first()
        if not reg:
            return Response({"registered": False})
        return Response({
            "registered": True,
            "id": reg.id,
            "player_id": reg.player_id,
            "payment_status": reg.payment_status,
            "event_password_required": bool(quiz.event_password)
        })


class QuizAttemptStartView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        reg = get_object_or_404(QuizRegistration, quiz=quiz, student=request.user)
        if reg.payment_status != 'paid':
            return Response({"detail": "Payment required."}, status=403)
            
        attempt, created = QuizAttempt.objects.get_or_create(student=request.user, quiz=quiz)
        return Response({"attempt_id": attempt.id, "current_index": attempt.current_question_index})


class QuizAttemptNextQuestionView(APIView):
    permission_classes = [IsStudentUser]
    
    def get(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        attempt = get_object_or_404(QuizAttempt, student=request.user, quiz=quiz)
        
        questions = list(quiz.questions.filter(question_type=Question.QuestionType.REGULAR).order_by('order', 'id'))
        if attempt.current_question_index >= len(questions):
            return Response({"completed": True})
            
        question = questions[attempt.current_question_index]
        choices = [{"id": c.id, "text": c.text} for c in question.choices.all()]
        
        return Response({
            "completed": False,
            "question": {
                "id": question.id,
                "text": question.text,
                "marks": question.marks,
                "order": question.order,
                "choices": choices,
                "total_questions": len(questions),
                "current_index": attempt.current_question_index
            }
        })


class QuizAttemptSubmitAnswerView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        attempt = get_object_or_404(QuizAttempt, student=request.user, quiz=quiz)
        
        questions = list(quiz.questions.filter(question_type=Question.QuestionType.REGULAR).order_by('order', 'id'))
        if attempt.current_question_index >= len(questions):
            return Response({"detail": "Quiz already completed."}, status=400)
            
        question = questions[attempt.current_question_index]
        choice_id = request.data.get('choice_id')
        time_taken = request.data.get('time_taken', 0)
        
        selected_choice = None
        if choice_id:
            selected_choice = Choice.objects.filter(question=question, id=choice_id).first()
            if selected_choice and selected_choice.is_correct:
                attempt.score += question.marks
                
        StudentAnswer.objects.create(
            attempt=attempt,
            question=question,
            selected_choice=selected_choice,
            time_taken_seconds=time_taken
        )
        
        attempt.current_question_index += 1
        attempt.save()
        
        is_completed = attempt.current_question_index >= len(questions)
        if is_completed:
            from django.utils import timezone
            attempt.completed_at = timezone.now()
            attempt.save()
            
        correct_choice = Choice.objects.filter(question=question, is_correct=True).first()
        correct_choice_data = None
        if correct_choice:
            correct_choice_data = {
                "id": correct_choice.id,
                "text": correct_choice.text
            }
            
        return Response({
            "correct": selected_choice.is_correct if selected_choice else False,
            "completed": is_completed,
            "correct_choice": correct_choice_data,
            "trivia": question.trivia
        })


class AdminStatsView(APIView):
    """
    Returns platform-wide statistics for the admin dashboard.
    """
    permission_classes = [IsAdminUser]
    
    def get(self, request):
        total_students = User.objects.filter(role='student').count()
        total_quizzes = Quiz.objects.filter(is_archived=False).count()
        active_quizzes = Quiz.objects.filter(is_archived=False, visible_to_students=True).count()
        total_registrations = QuizRegistration.objects.count()
        
        return Response({
            "total_students": total_students,
            "total_quizzes": total_quizzes,
            "active_quizzes": active_quizzes,
            "total_registrations": total_registrations
        })


class PublishedQuizListView(APIView):
    """
    Lists all published, non-archived quizzes for students.
    """
    permission_classes = [IsStudentUser]
    
    def get(self, request):
        quizzes = Quiz.objects.annotate(
            registered_count=Count('registrations')
        ).filter(
            visible_to_students=True,
            is_archived=False
        )
        return Response(QuizSerializer(quizzes, many=True).data)


class QuizDetailView(APIView):
    """
    Detailed view of a single quiz, available to authenticated users.
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, pk):
        quiz = get_object_or_404(
            Quiz.objects.annotate(registered_count=Count('registrations')), 
            pk=pk
        )
        
        # Prevent students from viewing hidden/archived quizzes
        if request.user.role == 'student' and (not quiz.visible_to_students or quiz.is_archived):
            return Response({"detail": "Quiz not found or not available."}, status=status.HTTP_404_NOT_FOUND)
            
        return Response(QuizSerializer(quiz).data)


class StudentRegistrationView(APIView):
    """
    Handles student registration for a quiz.
    """
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        
        try:
            registration = register_student_for_quiz(request.user, quiz)
            return Response(
                QuizRegistrationSerializer(registration).data, 
                status=status.HTTP_201_CREATED
            )
        except DjangoValidationError as e:
            raise ValidationError({"detail": str(e.message) if hasattr(e, 'message') else str(e)})


class MockPaymentView(APIView):
    """
    Handles the simulated payment success for a pending registration.
    """
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        registration = get_object_or_404(
            QuizRegistration, 
            quiz_id=pk, 
            student=request.user
        )
        
        try:
            processed_registration = process_mock_payment(registration)
            return Response(QuizRegistrationSerializer(processed_registration).data)
        except DjangoValidationError as e:
            raise ValidationError({"detail": str(e.message) if hasattr(e, 'message') else str(e)})


class MyRegistrationsView(APIView):
    """
    Returns the authenticated student's registrations.
    """
    permission_classes = [IsStudentUser]
    
    def get(self, request):
        registrations = QuizRegistration.objects.select_related('quiz').filter(
            student=request.user
        )
        return Response(QuizRegistrationSerializer(registrations, many=True).data)

from quizzes.models import Team
from quizzes.serializers import TeamSerializer
from django.db.models import Q

class StudentTeamViewSet(viewsets.ModelViewSet):
    permission_classes = [IsStudentUser]
    serializer_class = TeamSerializer
    
    def get_queryset(self):
        # Teams for quizzes the user is registered for
        return Team.objects.filter(
            Q(leader=self.request.user) | Q(members=self.request.user) | Q(quiz__registrations__student=self.request.user)
        ).distinct()
        
    def perform_create(self, serializer):
        quiz_id = self.request.data.get('quiz')
        quiz = get_object_or_404(Quiz, pk=quiz_id)
        if not QuizRegistration.objects.filter(quiz=quiz, student=self.request.user).exists():
            raise ValidationError({"detail": "You must be registered for this quiz to create a team."})
        serializer.save(leader=self.request.user, quiz=quiz)
        
    @action(detail=True, methods=['post'])
    def join(self, request, pk=None):
        team = self.get_object()
        if not QuizRegistration.objects.filter(quiz=team.quiz, student=request.user).exists():
            return Response({"detail": "You must be registered for this quiz to join this team."}, status=400)
            
        team.members.add(request.user)
        return Response({"detail": "Successfully joined team!"})


# Student KBC Event views
class VerifyQuizAccessView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        player_id = request.data.get('player_id', '').strip()
        password = request.data.get('event_password', '').strip()
        
        reg = QuizRegistration.objects.filter(quiz=quiz, student=request.user).first()
        if not reg or reg.payment_status != 'paid':
            return Response({"detail": "You are not registered or paid for this quiz."}, status=403)
            
        if reg.player_id.lower() != player_id.lower():
            return Response({"detail": "Invalid Player ID."}, status=400)
            
        if quiz.event_password and quiz.event_password != password:
            return Response({"detail": "Invalid Event Password. Please ask the organizers."}, status=400)
            
        return Response({"success": True, "detail": "Access granted to the arena."})


class QuizLiveStateView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        
        # Auto-populate batches if empty and stage is batch_selection or later FFF/Hotseat stages
        if quiz.current_stage != Quiz.Stage.REGULAR and not quiz.batch_1_players and not quiz.batch_2_players and not quiz.batch_3_players:
            attempts = list(QuizAttempt.objects.filter(quiz=quiz, completed_at__isnull=False).order_by('-score', 'completed_at'))
            if attempts:
                student_ids = [att.student_id for att in attempts]
                top_30 = student_ids[:30]
                quiz.batch_1_players = top_30[0:10]
                quiz.batch_2_players = top_30[10:20]
                quiz.batch_3_players = top_30[20:30]
                quiz.top_30_selected = top_30
                quiz.save(update_fields=['top_30_selected', 'batch_1_players', 'batch_2_players', 'batch_3_players'])

        user = request.user
        
        role = "spectator"
        batch_number = None
        is_in_active_batch = False
        hotseat_attempt_data = None
        fff_question_data = None
        fff_answered = False
        
        if user.role == 'student':
            user_id = user.id
            is_testing_quiz = "test" in quiz.title.lower()
            
            if user_id in quiz.batch_1_players:
                role = "batch_player"
                batch_number = 1
            elif user_id in quiz.batch_2_players:
                role = "batch_player"
                batch_number = 2
            elif user_id in quiz.batch_3_players:
                role = "batch_player"
                batch_number = 3
            elif is_testing_quiz:
                role = "batch_player"
                stage = quiz.current_stage
                if stage in [Quiz.Stage.FFF_BATCH_2, Quiz.Stage.HOTSEAT_BATCH_2]:
                    batch_number = 2
                elif stage in [Quiz.Stage.FFF_BATCH_3, Quiz.Stage.HOTSEAT_BATCH_3]:
                    batch_number = 3
                else:
                    batch_number = 1
                
            stage = quiz.current_stage
            if (stage == Quiz.Stage.FFF_BATCH_1 and batch_number == 1) or \
               (stage == Quiz.Stage.FFF_BATCH_2 and batch_number == 2) or \
               (stage == Quiz.Stage.FFF_BATCH_3 and batch_number == 3):
                is_in_active_batch = True
                
            if (stage == Quiz.Stage.HOTSEAT_BATCH_1 and quiz.hotseat_player_1 == user) or \
               (stage == Quiz.Stage.HOTSEAT_BATCH_2 and quiz.hotseat_player_2 == user) or \
               (stage == Quiz.Stage.HOTSEAT_BATCH_3 and quiz.hotseat_player_3 == user):
                role = "hotseat_player"
                
            # Load FFF question if active
            fff_question_data = None
            fff_answered = False
            fff_q_type = None
            fff_batch = None
            if stage == Quiz.Stage.FFF_BATCH_1:
                fff_q_type = Question.QuestionType.FFF_1
                fff_batch = 1
            elif stage == Quiz.Stage.FFF_BATCH_2:
                fff_q_type = Question.QuestionType.FFF_2
                fff_batch = 2
            elif stage == Quiz.Stage.FFF_BATCH_3:
                fff_q_type = Question.QuestionType.FFF_3
                fff_batch = 3
                
            if fff_q_type:
                fff_q = Question.objects.filter(quiz=quiz, question_type=fff_q_type).first()
                if fff_q:
                    fff_question_data = {
                        "id": fff_q.id,
                        "text": fff_q.text,
                        "choices": [{"id": c.id, "text": c.text} for c in fff_q.choices.all()]
                    }
                    fff_answered = FFFAnswer.objects.filter(
                        quiz=quiz, student=user, batch_number=fff_batch, question=fff_q
                    ).exists()

        # Load active hotseat attempt data for all roles (students, hosts, and spectators)
        stage = quiz.current_stage
        active_hotseat_player = None
        active_batch = None
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            active_hotseat_player = quiz.hotseat_player_1
            active_batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            active_hotseat_player = quiz.hotseat_player_2
            active_batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            active_hotseat_player = quiz.hotseat_player_3
            active_batch = 3
            
        if active_hotseat_player:
            attempt = HotseatAttempt.objects.filter(quiz=quiz, student=active_hotseat_player, batch_number=active_batch).first()
            if attempt:
                hotseat_attempt_data = HotseatAttemptSerializer(attempt).data
                    
        live_participants = quiz.registrations.count()
        total_questions = quiz.questions.filter(question_type=Question.QuestionType.REGULAR).count()
        overall_total_questions = quiz.questions.count()
                    
        # Resolve batch player names in a single database query
        all_ids = set((quiz.batch_1_players or []) + (quiz.batch_2_players or []) + (quiz.batch_3_players or []))
        user_map = {}
        if all_ids:
            users = User.objects.filter(id__in=all_ids)
            user_map = {u.id: u.full_name for u in users}

        def resolve_players_list(id_list):
            if not id_list:
                return []
            return [{"id": pid, "name": user_map.get(pid, f"Player ID: {pid}")} for pid in id_list]

        b1_resolved = resolve_players_list(quiz.batch_1_players)
        b2_resolved = resolve_players_list(quiz.batch_2_players)
        b3_resolved = resolve_players_list(quiz.batch_3_players)

        prefs = SystemPreferences.get_solo()

        return Response({
            "quiz_id": quiz.id,
            "title": quiz.title,
            "intro_title": quiz.intro_title or "Kaun Banega Codepati",
            "current_stage": quiz.current_stage,
            "student_role": role,
            "batch_number": batch_number,
            "is_in_active_batch": is_in_active_batch,
            "hotseat_attempt": hotseat_attempt_data,
            "stage_display": quiz.get_current_stage_display(),
            "fff_question": fff_question_data,
            "fff_answered": fff_answered,
            "live_participants": live_participants,
            "total_questions": total_questions,
            "overall_total_questions": overall_total_questions,
            "batch_1_players": b1_resolved,
            "batch_2_players": b2_resolved,
            "batch_3_players": b3_resolved,
            "prelim_mcq_timer": prefs.prelim_mcq_timer,
            "fff_speed_timer": prefs.fff_speed_timer,
            "hotseat_q1_q5_limit": prefs.hotseat_q1_q5_limit,
            "hotseat_q6_q10_limit": prefs.hotseat_q6_q10_limit,
            "auto_approve_registrations": prefs.auto_approve_registrations,
        })


class FFFSubmitView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.FFF_BATCH_1:
            batch_num = 1
            q_type = Question.QuestionType.FFF_1
            players = quiz.batch_1_players
        elif stage == Quiz.Stage.FFF_BATCH_2:
            batch_num = 2
            q_type = Question.QuestionType.FFF_2
            players = quiz.batch_2_players
        elif stage == Quiz.Stage.FFF_BATCH_3:
            batch_num = 3
            q_type = Question.QuestionType.FFF_3
            players = quiz.batch_3_players
        else:
            return Response({"detail": "Fastest Finger First is not active at this stage."}, status=400)
            
        is_testing_quiz = "test" in quiz.title.lower()
        if request.user.id not in players and not is_testing_quiz:
            return Response({"detail": "You are not in the active batch for this Fastest Finger First round."}, status=403)
            
        fff_question = Question.objects.filter(quiz=quiz, question_type=q_type).first()
        if not fff_question:
            return Response({"detail": "FFF question not found."}, status=404)
            
        if FFFAnswer.objects.filter(quiz=quiz, student=request.user, batch_number=batch_num, question=fff_question).exists():
            return Response({"detail": "You have already submitted your answer for this round."}, status=400)
            
        selected_sequence = request.data.get('selected_sequence', [])
        time_taken = float(request.data.get('time_taken', 0.0))
        
        # If selected_sequence is not sent but choice_id is, wrap it in a list
        choice_id = request.data.get('choice_id')
        if not selected_sequence and choice_id:
            selected_sequence = [choice_id]
            
        correct_choices = Choice.objects.filter(question=fff_question, correct_order__isnull=False).order_by('correct_order')
        
        if correct_choices.exists():
            correct_sequence = [c.id for c in correct_choices]
            try:
                student_seq = [int(x) for x in selected_sequence]
            except (ValueError, TypeError):
                student_seq = []
            is_sequence_correct = (student_seq == correct_sequence)
        else:
            # Fallback to old single-choice logic
            first_choice_id = selected_sequence[0] if selected_sequence else None
            selected_choice = None
            if first_choice_id:
                selected_choice = Choice.objects.filter(question=fff_question, id=first_choice_id).first()
            is_sequence_correct = selected_choice.is_correct if selected_choice else False
        
        first_choice = None
        if selected_sequence:
            try:
                first_choice_id = int(selected_sequence[0])
                first_choice = Choice.objects.filter(question=fff_question, id=first_choice_id).first()
            except (ValueError, TypeError, IndexError):
                pass
        
        answer = FFFAnswer.objects.create(
            quiz=quiz,
            student=request.user,
            batch_number=batch_num,
            question=fff_question,
            selected_choice=first_choice,
            time_taken_seconds=time_taken,
            is_correct=is_sequence_correct,
            submitted_sequence=",".join(map(str, selected_sequence))
        )
        
        return Response({
            "submitted": True,
            "correct": is_sequence_correct
        })


class HotseatQuestionView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
            q_type = Question.QuestionType.HOTSEAT_1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
            q_type = Question.QuestionType.HOTSEAT_2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
            q_type = Question.QuestionType.HOTSEAT_3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not hotseat_player:
            return Response({"detail": "No hotseat contestant has been promoted yet."}, status=404)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=hotseat_player, batch_number=batch_num)
        
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"completed": True, "status": attempt.status, "score": attempt.score})
            
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        if not questions:
            return Response({"detail": "No hotseat questions have been uploaded for this batch."}, status=404)
            
        if attempt.current_question_index >= len(questions):
            attempt.status = HotseatAttempt.Status.COMPLETED
            attempt.completed_at = timezone.now()
            attempt.save()
            return Response({"completed": True, "status": attempt.status, "score": attempt.score})
            
        question = questions[attempt.current_question_index]
        choices = list(question.choices.all())
        
        return Response({
            "completed": False,
            "current_index": attempt.current_question_index,
            "total_questions": len(questions),
            "score": attempt.score,
            "preselected_choice_id": attempt.preselected_choice_id,
            "question": {
                "id": question.id,
                "text": question.text,
                "category": question.category,
                "order": question.order,
                "choices": [{"id": c.id, "text": c.text} for c in choices]
            }
        })


class HotseatPreselectView(APIView):
    """Allows the hotseat player to save their selected choice without locking it."""
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
        else:
            return Response({"detail": "Hotseat is not active."}, status=400)
        
        if request.user != hotseat_player:
            return Response({"detail": "You are not the hotseat contestant."}, status=403)
        
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=request.user, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt is not active."}, status=400)
        
        choice_id = request.data.get('choice_id')
        if choice_id:
            choice = Choice.objects.filter(id=choice_id).first()
            attempt.preselected_choice = choice
        else:
            attempt.preselected_choice = None
        attempt.save(update_fields=['preselected_choice'])
        
        return Response({"detail": "Selection updated.", "preselected_choice_id": choice_id})


SCORE_LADDER = [
    10,   # Q1
    20,   # Q2
    30,   # Q3
    40,   # Q4
    50,   # Q5 (Checkpoint 1)
    60,   # Q6
    70,   # Q7
    80,   # Q8
    90,   # Q9
    100,  # Q10 (Checkpoint 2)
    110,  # Q11
    120,  # Q12
    130,  # Q13
    140,  # Q14
    150   # Q15 (Checkpoint 3)
]

class HotseatSubmitView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
            q_type = Question.QuestionType.HOTSEAT_1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
            q_type = Question.QuestionType.HOTSEAT_2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
            q_type = Question.QuestionType.HOTSEAT_3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if request.user != hotseat_player:
            return Response({"detail": "You are not the active hotseat contestant."}, status=403)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=request.user, batch_number=batch_num)
        
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        if attempt.current_question_index >= len(questions):
            return Response({"detail": "All questions completed."}, status=400)
            
        question = questions[attempt.current_question_index]
        choice_id = request.data.get('choice_id')
        
        selected_choice = Choice.objects.filter(question=question, id=choice_id).first()
        is_correct = selected_choice.is_correct if selected_choice else False
        
        if is_correct:
            current_points = SCORE_LADDER[attempt.current_question_index]
            attempt.score = current_points
            attempt.current_question_index += 1
            attempt.current_question_switched = False
            
            if attempt.current_question_index >= len(questions):
                attempt.status = HotseatAttempt.Status.COMPLETED
                attempt.completed_at = timezone.now()
                self.save_quiz_hotseat_score(quiz, batch_num, attempt.score, "completed")
            else:
                self.save_quiz_hotseat_score(quiz, batch_num, attempt.score, "playing")
                
            attempt.save()
            return Response({
                "correct": True,
                "current_points": attempt.score,
                "next_index": attempt.current_question_index,
                "completed": attempt.status == HotseatAttempt.Status.COMPLETED
            })
        else:
            checkpoint_score = 0
            fail_index = attempt.current_question_index
            if fail_index >= 10:
                checkpoint_score = 100  # Drop to Checkpoint 2 score
            elif fail_index >= 5:
                checkpoint_score = 50   # Drop to Checkpoint 1 score
            else:
                checkpoint_score = 0
                
            attempt.score = checkpoint_score
            attempt.status = HotseatAttempt.Status.FAILED
            attempt.completed_at = timezone.now()
            attempt.save()
            
            self.save_quiz_hotseat_score(quiz, batch_num, attempt.score, "failed")
            
            return Response({
                "correct": False,
                "correct_choice_id": Choice.objects.filter(question=question, is_correct=True).values_list('id', flat=True).first(),
                "checkpoint_points": attempt.score,
                "completed": True
            })
            
    def save_quiz_hotseat_score(self, quiz, batch_num, score, status):
        if batch_num == 1:
            quiz.hotseat_score_1 = score
            quiz.hotseat_status_1 = status
            quiz.save(update_fields=['hotseat_score_1', 'hotseat_status_1'])
        elif batch_num == 2:
            quiz.hotseat_score_2 = score
            quiz.hotseat_status_2 = status
            quiz.save(update_fields=['hotseat_score_2', 'hotseat_status_2'])
        elif batch_num == 3:
            quiz.hotseat_score_3 = score
            quiz.hotseat_status_3 = status
            quiz.save(update_fields=['hotseat_score_3', 'hotseat_status_3'])


class HotseatLifelineView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
            q_type = Question.QuestionType.HOTSEAT_1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
            q_type = Question.QuestionType.HOTSEAT_2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
            q_type = Question.QuestionType.HOTSEAT_3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if request.user != hotseat_player:
            return Response({"detail": "You are not the active hotseat contestant."}, status=403)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=request.user, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        if not attempt.options_visible:
            return Response({"detail": "Lifelines can only be requested after choices are shown by the host."}, status=400)
            
        if attempt.current_question_index >= 10:
            return Response({"detail": "Lifelines are no longer available after the 10th question."}, status=400)
            
        lifeline = request.data.get('lifeline') or request.data.get('lifeline_type')
        if not lifeline or lifeline not in ['5050', 'poll', 'switch']:
            return Response({"detail": "Invalid lifeline provided."}, status=400)
            
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        question = questions[attempt.current_question_index]
        choices = list(question.choices.all())
        
        if lifeline == '5050':
            if attempt.lifeline_5050_used:
                return Response({"detail": "50:50 lifeline already used."}, status=400)
                
            correct_choice = next(c for c in choices if c.is_correct)
            incorrect_choices = [c for c in choices if not c.is_correct]
            
            random.shuffle(incorrect_choices)
            eliminated = incorrect_choices[:2]
            
            attempt.lifeline_5050_used = True
            attempt.save()
            
            return Response({
                "lifeline": "5050",
                "eliminated_choice_ids": [c.id for c in eliminated]
            })
            
        elif lifeline == 'poll':
            if attempt.lifeline_poll_used:
                return Response({"detail": "Audience Poll lifeline already used."}, status=400)
                
            correct_choice = next(c for c in choices if c.is_correct)
            
            correct_votes = random.randint(55, 75)
            remaining_votes = 100 - correct_votes
            
            incorrect_choices = [c for c in choices if not c.is_correct]
            random.shuffle(incorrect_choices)
            
            poll_results = {}
            poll_results[correct_choice.id] = correct_votes
            
            if len(incorrect_choices) >= 3:
                v1 = random.randint(5, max(5, remaining_votes - 10))
                remaining_votes -= v1
                v2 = random.randint(2, max(2, remaining_votes - 5))
                remaining_votes -= v2
                v3 = remaining_votes
                
                poll_results[incorrect_choices[0].id] = v1
                poll_results[incorrect_choices[1].id] = v2
                poll_results[incorrect_choices[2].id] = v3
            else:
                for idx, inc in enumerate(incorrect_choices):
                    if idx == len(incorrect_choices) - 1:
                        poll_results[inc.id] = remaining_votes
                    else:
                        v = random.randint(5, max(5, remaining_votes // 2))
                        poll_results[inc.id] = v
                        remaining_votes -= v
                        
            attempt.lifeline_poll_used = True
            attempt.save()
            
            return Response({
                "lifeline": "poll",
                "votes": poll_results
            })
            
        elif lifeline == 'switch':
            if attempt.lifeline_switch_used:
                return Response({"detail": "Switch Question lifeline already used."}, status=400)
                
            category = request.data.get('category', 'General')
            
            replacement = Question.objects.filter(
                quiz=quiz,
                question_type=q_type,
                category__iexact=category
            ).exclude(id=question.id).first()
            
            if not replacement:
                replacement = Question.objects.filter(
                    quiz=quiz,
                    question_type=q_type
                ).exclude(id=question.id).first()
                
            if not replacement:
                return Response({"detail": "No replacement questions available."}, status=404)
                
            original_order = question.order
            question.order = replacement.order
            replacement.order = original_order
            
            question.save(update_fields=['order'])
            replacement.save(update_fields=['order'])
            
            attempt.lifeline_switch_used = True
            attempt.save()
            
            choices = list(replacement.choices.all())
            return Response({
                "lifeline": "switch",
                "question": {
                    "id": replacement.id,
                    "text": replacement.text,
                    "category": replacement.category,
                    "order": original_order,
                    "choices": [{"id": c.id, "text": c.text} for c in choices]
                }
            })


class HotseatWalkAwayView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if request.user != hotseat_player:
            return Response({"detail": "You are not the active hotseat contestant."}, status=403)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=request.user, batch_number=batch_num)
        
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        attempt.status = HotseatAttempt.Status.WALKED_AWAY
        attempt.completed_at = timezone.now()
        attempt.save()
        
        if batch_num == 1:
            quiz.hotseat_score_1 = attempt.score
            quiz.hotseat_status_1 = "walked_away"
            quiz.save(update_fields=['hotseat_score_1', 'hotseat_status_1'])
        elif batch_num == 2:
            quiz.hotseat_score_2 = attempt.score
            quiz.hotseat_status_2 = "walked_away"
            quiz.save(update_fields=['hotseat_score_2', 'hotseat_status_2'])
        elif batch_num == 3:
            quiz.hotseat_score_3 = attempt.score
            quiz.hotseat_status_3 = "walked_away"
            quiz.save(update_fields=['hotseat_score_3', 'hotseat_status_3'])
            
        return Response({
            "walked_away": True,
            "final_points": attempt.score
        })


class HotseatLifelineRequestView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if request.user != hotseat_player:
            return Response({"detail": "You are not the active hotseat contestant."}, status=403)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=request.user, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        if not attempt.options_visible:
            return Response({"detail": "Lifelines can only be requested after choices are shown by the host."}, status=400)
            
        if attempt.current_question_index >= 10:
            return Response({"detail": "Lifelines are no longer available after the 10th question."}, status=400)
            
        lifeline = request.data.get('lifeline') or request.data.get('lifeline_type')
        category = request.data.get('category', '')
        
        if not lifeline or lifeline not in ['5050', 'poll', 'switch']:
            return Response({"detail": "Invalid lifeline provided."}, status=400)
            
        if lifeline == '5050' and attempt.lifeline_5050_used:
            return Response({"detail": "50:50 lifeline already used."}, status=400)
        elif lifeline == 'poll' and attempt.lifeline_poll_used:
            return Response({"detail": "Audience Poll lifeline already used."}, status=400)
        elif lifeline == 'switch' and attempt.lifeline_switch_used:
            return Response({"detail": "Switch Question lifeline already used."}, status=400)
            
        attempt.pending_lifeline_type = lifeline
        attempt.pending_lifeline_switch_category = category
        attempt.lifeline_request_status = 'requested'
        attempt.approved_lifeline_data = {}
        attempt.save()
        
        return Response({
            "detail": f"Request for {lifeline} lifeline submitted to host.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class HotseatLifelineAcknowledgeView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if request.user != hotseat_player:
            return Response({"detail": "You are not the active hotseat contestant."}, status=403)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=request.user, batch_number=batch_num)
        
        attempt.lifeline_request_status = 'none'
        attempt.pending_lifeline_type = ''
        attempt.pending_lifeline_switch_category = ''
        attempt.save()
        
        return Response({
            "detail": "Lifeline status acknowledged.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminApproveLifelineView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
            q_type = Question.QuestionType.HOTSEAT_1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
            q_type = Question.QuestionType.HOTSEAT_2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
            q_type = Question.QuestionType.HOTSEAT_3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not hotseat_player:
            return Response({"detail": "No active hotseat contestant."}, status=404)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=hotseat_player, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        if attempt.lifeline_request_status != 'requested':
            return Response({"detail": "No pending lifeline request to approve."}, status=400)
            
        lifeline = attempt.pending_lifeline_type
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        question = questions[attempt.current_question_index]
        choices = list(question.choices.all())
        
        approved_data = {}
        
        if lifeline == '5050':
            if attempt.lifeline_5050_used:
                return Response({"detail": "50:50 lifeline already used."}, status=400)
                
            correct_choice = next(c for c in choices if c.is_correct)
            incorrect_choices = [c for c in choices if not c.is_correct]
            
            import random
            random.shuffle(incorrect_choices)
            eliminated = incorrect_choices[:2]
            
            attempt.lifeline_5050_used = True
            approved_data = {
                "eliminated_choice_ids": [c.id for c in eliminated]
            }
            
        elif lifeline == 'poll':
            if attempt.lifeline_poll_used:
                return Response({"detail": "Audience Poll lifeline already used."}, status=400)
                
            correct_choice = next(c for c in choices if c.is_correct)
            
            import random
            correct_votes = random.randint(55, 75)
            remaining_votes = 100 - correct_votes
            
            incorrect_choices = [c for c in choices if not c.is_correct]
            random.shuffle(incorrect_choices)
            
            poll_results = {}
            poll_results[correct_choice.id] = correct_votes
            
            if len(incorrect_choices) >= 3:
                v1 = random.randint(5, max(5, remaining_votes - 10))
                remaining_votes -= v1
                v2 = random.randint(2, max(2, remaining_votes - 5))
                remaining_votes -= v2
                v3 = remaining_votes
                
                poll_results[incorrect_choices[0].id] = v1
                poll_results[incorrect_choices[1].id] = v2
                poll_results[incorrect_choices[2].id] = v3
            else:
                for idx, inc in enumerate(incorrect_choices):
                    if idx == len(incorrect_choices) - 1:
                        poll_results[inc.id] = remaining_votes
                    else:
                        v = random.randint(5, max(5, remaining_votes // 2))
                        poll_results[inc.id] = v
                        remaining_votes -= v
                        
            attempt.lifeline_poll_used = True
            approved_data = {
                "votes": poll_results
            }
            
        elif lifeline == 'switch':
            if attempt.lifeline_switch_used:
                return Response({"detail": "Switch Question lifeline already used."}, status=400)
            approved_data = {}
            
        attempt.lifeline_request_status = 'approved'
        attempt.approved_lifeline_data = approved_data
        attempt.save()
        
        return Response({
            "detail": f"{lifeline} lifeline approved successfully.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminRejectLifelineView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not hotseat_player:
            return Response({"detail": "No active hotseat player promoted."}, status=400)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=hotseat_player, batch_number=batch_num)
        
        if attempt.lifeline_request_status != 'requested':
            return Response({"detail": "No pending lifeline request to reject."}, status=400)
            
        attempt.lifeline_request_status = 'rejected'
        attempt.save()
        
        return Response({
            "detail": "Lifeline request rejected successfully.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminShowOptionsView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            player = quiz.hotseat_player_1
            batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            player = quiz.hotseat_player_2
            batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            player = quiz.hotseat_player_3
            batch = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not player:
            return Response({"detail": "No active hotseat player promoted."}, status=400)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=player, batch_number=batch)
        attempt.options_visible = True
        attempt.timer_is_paused = False
        attempt.save()
        
        return Response({
            "detail": "Choices successfully revealed.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminPauseTimerView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            player = quiz.hotseat_player_1
            batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            player = quiz.hotseat_player_2
            batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            player = quiz.hotseat_player_3
            batch = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not player:
            return Response({"detail": "No active hotseat player promoted."}, status=400)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=player, batch_number=batch)
        attempt.timer_is_paused = True
        attempt.save()
        
        return Response({
            "detail": "Timer paused.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminResumeTimerView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            player = quiz.hotseat_player_1
            batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            player = quiz.hotseat_player_2
            batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            player = quiz.hotseat_player_3
            batch = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not player:
            return Response({"detail": "No active hotseat player promoted."}, status=400)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=player, batch_number=batch)
        attempt.timer_is_paused = False
        attempt.save()
        
        return Response({
            "detail": "Timer resumed.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminNextQuestionReadyView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            player = quiz.hotseat_player_1
            batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            player = quiz.hotseat_player_2
            batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            player = quiz.hotseat_player_3
            batch = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not player:
            return Response({"detail": "No active hotseat player promoted."}, status=400)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=player, batch_number=batch)
        attempt.showing_question = True
        attempt.options_visible = False
        attempt.timer_is_paused = False
        attempt.save()
        
        return Response({
            "detail": "Next question pushed to contestant.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminTriggerIntroView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            player = quiz.hotseat_player_1
            batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            player = quiz.hotseat_player_2
            batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            player = quiz.hotseat_player_3
            batch = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not player:
            return Response({"detail": "No active hotseat player promoted."}, status=400)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=player, batch_number=batch)
        attempt.show_intro = True
        attempt.save()
        
        return Response({
            "detail": "KBC Intro playback triggered successfully.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminCompleteIntroView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            player = quiz.hotseat_player_1
            batch = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            player = quiz.hotseat_player_2
            batch = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            player = quiz.hotseat_player_3
            batch = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not player:
            return Response({"detail": "No active hotseat player promoted."}, status=400)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=player, batch_number=batch)
        attempt.show_intro = False
        attempt.intro_played = True
        attempt.save()
        
        return Response({
            "detail": "KBC Intro playback completed.",
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class StudentSwitchCategoryListView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        categories = quiz.switch_categories.all().select_related('question')
        categories = [c for c in categories if c.question]
        
        data = []
        for c in categories:
            img_url = c.image.url if c.image else None
            if img_url and request:
                img_url = request.build_absolute_uri(img_url)
            data.append({
                "id": c.id,
                "name": c.name,
                "image": img_url
            })
        return Response(data)


class HotseatSelectSwitchCategoryView(APIView):
    permission_classes = [IsStudentUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if request.user != hotseat_player:
            return Response({"detail": "You are not the active hotseat contestant."}, status=403)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=request.user, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        if attempt.lifeline_request_status != 'approved' or attempt.pending_lifeline_type != 'switch':
            return Response({"detail": "Switch Question lifeline has not been approved by the host."}, status=400)
            
        category_id = request.data.get('category_id')
        if not category_id:
            return Response({"detail": "Category ID is required."}, status=400)
            
        switch_cat = get_object_or_404(SwitchCategory, quiz=quiz, id=category_id)
        if not switch_cat.question:
            return Response({"detail": "No question configured for this category."}, status=400)
            
        with transaction.atomic():
            attempt.pending_lifeline_switch_category = f"{switch_cat.id}:{switch_cat.name}"
            attempt.save()
            
        return Response({
            "selected": True,
            "category_name": switch_cat.name,
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class AdminConfirmSwitchLifelineView(APIView):
    permission_classes = [IsAdminUser]
    
    def post(self, request, pk):
        quiz = get_object_or_404(Quiz, pk=pk)
        stage = quiz.current_stage
        
        if stage == Quiz.Stage.HOTSEAT_BATCH_1:
            hotseat_player = quiz.hotseat_player_1
            batch_num = 1
            q_type = Question.QuestionType.HOTSEAT_1
        elif stage == Quiz.Stage.HOTSEAT_BATCH_2:
            hotseat_player = quiz.hotseat_player_2
            batch_num = 2
            q_type = Question.QuestionType.HOTSEAT_2
        elif stage == Quiz.Stage.HOTSEAT_BATCH_3:
            hotseat_player = quiz.hotseat_player_3
            batch_num = 3
            q_type = Question.QuestionType.HOTSEAT_3
        else:
            return Response({"detail": "Hotseat is not active at this stage."}, status=400)
            
        if not hotseat_player:
            return Response({"detail": "No active hotseat contestant."}, status=404)
            
        attempt = get_object_or_404(HotseatAttempt, quiz=quiz, student=hotseat_player, batch_number=batch_num)
        if attempt.status != HotseatAttempt.Status.PLAYING:
            return Response({"detail": "Hotseat attempt already completed."}, status=400)
            
        if attempt.lifeline_request_status != 'approved' or attempt.pending_lifeline_type != 'switch':
            return Response({"detail": "Switch Question lifeline is not active."}, status=400)
            
        selected_cat_str = attempt.pending_lifeline_switch_category
        if not selected_cat_str or ":" not in selected_cat_str:
            return Response({"detail": "Contestant has not selected a switch category yet."}, status=400)
            
        try:
            category_id = int(selected_cat_str.split(":", 1)[0])
        except Exception:
            return Response({"detail": "Invalid selected category format."}, status=400)
            
        switch_cat = get_object_or_404(SwitchCategory, quiz=quiz, id=category_id)
        replacement_question = switch_cat.question
        if not replacement_question:
            return Response({"detail": "No question configured for this category."}, status=400)
            
        questions = list(Question.objects.filter(quiz=quiz, question_type=q_type).order_by('order', 'id'))
        original_question = questions[attempt.current_question_index]
        
        with transaction.atomic():
            original_order = original_question.order
            
            replacement_question.question_type = q_type
            replacement_question.order = original_order
            replacement_question.save()
            
            original_question.question_type = Question.QuestionType.SWITCH
            original_question.order = 999
            original_question.save()
            
            switch_cat.question = None
            switch_cat.save()
            
            attempt.lifeline_switch_used = True
            attempt.current_question_switched = True
            attempt.lifeline_request_status = 'none'
            attempt.pending_lifeline_type = ''
            attempt.pending_lifeline_switch_category = ''
            attempt.showing_question = True
            attempt.options_visible = False
            attempt.timer_is_paused = False
            attempt.save()
            
        return Response({
            "switched": True,
            "attempt": HotseatAttemptSerializer(attempt).data
        })


class SystemPreferencesView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        prefs = SystemPreferences.get_solo()
        serializer = SystemPreferencesSerializer(prefs)
        return Response(serializer.data)

    def post(self, request):
        prefs = SystemPreferences.get_solo()
        serializer = SystemPreferencesSerializer(prefs, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)




