from django.utils import timezone
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Branch, Program, School, StudentProfile, User
from .serializers import (
    AdminStudentCreateSerializer,
    AdminStudentUpdateSerializer,
    BranchSerializer,
    LoginSerializer,
    ProgramSerializer,
    SchoolSerializer,
    StudentPublicSerializer,
    UserPublicSerializer,
)


from .authentication import get_user_from_request_token


class SchoolListView(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, _request):
        schools = School.objects.all()
        return Response(SchoolSerializer(schools, many=True).data)


class ProgramListView(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request):
        school_id = request.query_params.get("school_id") or request.query_params.get("school")
        programs = Program.objects.select_related("school")

        if school_id:
            programs = programs.filter(school_id=school_id)

        return Response(ProgramSerializer(programs, many=True).data)


class BranchListView(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request):
        program_id = request.query_params.get("program_id") or request.query_params.get("program")
        branches = Branch.objects.select_related("program")

        if program_id:
            branches = branches.filter(program_id=program_id)

        return Response(BranchSerializer(branches, many=True).data)





class LoginView(APIView):
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        user = serializer.validated_data["user"]
        user.rotate_session_token()
        user.last_login_at = timezone.now()
        user.save(update_fields=["session_token", "last_login_at", "updated_at"])

        return Response(
            {
                "message": "Login successful.",
                "token": str(user.session_token),
                "role": user.role,
                "user": UserPublicSerializer(user).data,
                "student": StudentPublicSerializer(user).data,
            }
        )


class CurrentUserView(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request):
        user = get_user_from_request_token(request)

        if user is None:
            return Response({"detail": "Authentication credentials were not provided or are invalid."}, status=401)

        return Response(
            {
                "role": user.role,
                "user": UserPublicSerializer(user).data,
            }
        )

    def put(self, request):
        user = get_user_from_request_token(request)

        if user is None:
            return Response({"detail": "Authentication credentials were not provided or are invalid."}, status=401)

        email = request.data.get("email", "").strip()
        password = request.data.get("password", "").strip()
        current_password = request.data.get("current_password", "").strip()

        if email:
            if User.objects.exclude(id=user.id).filter(email=email).exists():
                return Response({"detail": "This email is already taken by another account."}, status=status.HTTP_400_BAD_REQUEST)
            user.email = email

        if password:
            if not current_password:
                return Response({"detail": "Current password is required to change password."}, status=status.HTTP_400_BAD_REQUEST)
            if not user.check_password(current_password):
                return Response({"detail": "Current password is incorrect."}, status=status.HTTP_400_BAD_REQUEST)
            if len(password) < 8:
                return Response({"detail": "New password must be at least 8 characters long."}, status=status.HTTP_400_BAD_REQUEST)
            user.set_password(password)

        user.save()
        return Response(
            {
                "message": "Account credentials updated successfully.",
                "user": UserPublicSerializer(user).data,
            }
        )


import io
import openpyxl
from django.http import HttpResponse


class AdminStudentListView(APIView):
    authentication_classes = []
    permission_classes = []

    def _require_admin(self, request):
        user = get_user_from_request_token(request)
        if user is None or user.role != User.Role.ADMIN:
            return None
        return user

    def get(self, request):
        admin_user = self._require_admin(request)
        if not admin_user:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        students = User.objects.filter(role=User.Role.STUDENT).select_related(
            "student_profile__school",
            "student_profile__program",
            "student_profile__branch",
        ).order_by("-created_at")

        # Enforce school boundary for school admins
        if not admin_user.is_super_admin:
            if admin_user.school:
                students = students.filter(student_profile__school=admin_user.school)
            else:
                students = students.none()
        else:
            # Super admin can filter by school
            school_filter = request.query_params.get("school_id") or request.query_params.get("school")
            if school_filter:
                students = students.filter(student_profile__school_id=school_filter)

        # Both super admin and school admin can filter by program and branch
        program_filter = request.query_params.get("program_id") or request.query_params.get("program")
        if program_filter:
            students = students.filter(student_profile__program_id=program_filter)

        branch_filter = request.query_params.get("branch_id") or request.query_params.get("branch")
        if branch_filter:
            students = students.filter(student_profile__branch_id=branch_filter)

        data = []
        for s in students:
            profile = getattr(s, "student_profile", None)
            data.append({
                "id": s.id,
                "full_name": s.full_name,
                "roll_number": s.roll_number or "",
                "college_id": s.college_id,
                "email": s.email,
                "school": profile.school.school_code if profile else "",
                "school_id": profile.school.id if profile else None,
                "school_name": profile.school.school_name if profile else "",
                "program": profile.program.program_code if profile else "",
                "program_id": profile.program.id if profile else None,
                "program_name": profile.program.program_name if profile else "",
                "branch": profile.branch.branch_code if profile else "",
                "branch_id": profile.branch.id if profile else None,
                "branch_name": profile.branch.branch_name if profile else "",
                "year": profile.year if profile else "",
                "is_active": s.is_active,
                "created_at": s.created_at.isoformat(),
            })

        return Response(data)

    def post(self, request):
        if not self._require_admin(request):
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        serializer = AdminStudentCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()

        return Response(
            {"detail": "Student created successfully.", "student_id": user.id},
            status=status.HTTP_201_CREATED,
        )


class AdminStudentDetailView(APIView):
    authentication_classes = []
    permission_classes = []

    def put(self, request, pk):
        user = get_user_from_request_token(request)
        if user is None or user.role != User.Role.ADMIN:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        try:
            student = User.objects.get(pk=pk, role=User.Role.STUDENT)
        except User.DoesNotExist:
            return Response({"detail": "Student not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = AdminStudentUpdateSerializer(student, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response({"detail": "Student updated successfully."}, status=status.HTTP_200_OK)

    def delete(self, request, pk):
        user = get_user_from_request_token(request)
        if user is None or user.role != User.Role.ADMIN:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        try:
            student = User.objects.get(pk=pk, role=User.Role.STUDENT)
        except User.DoesNotExist:
            return Response({"detail": "Student not found."}, status=status.HTTP_404_NOT_FOUND)

        student.delete()
        return Response({"detail": "Student deleted successfully."}, status=status.HTTP_200_OK)


class AdminStudentBulkUploadView(APIView):
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        user = get_user_from_request_token(request)
        if user is None or user.role != User.Role.ADMIN:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        upload_file = request.FILES.get("file")
        if not upload_file:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(upload_file, read_only=True)
            ws = wb.active
        except Exception:
            return Response({"detail": "Invalid Excel file."}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created = 0
        skipped = 0
        errors = []

        from django.db import transaction as db_transaction

        for idx, row in enumerate(rows, start=2):
            if not row or len(row) < 6:
                skipped += 1
                continue

            full_name = str(row[0] or "").strip()
            roll_number = str(row[1] or "").strip().upper()
            college_id = roll_number  # Consolidate college ID and roll number
            email = str(row[2] or "").strip().lower()
            school_code = str(row[3] or "").strip()
            program_code = str(row[4] or "").strip()
            branch_code = str(row[5] or "").strip()
            year = str(row[6] if len(row) > 6 and row[6] else "1").strip()

            if not all([full_name, roll_number, email, school_code, program_code, branch_code]):
                errors.append(f"Row {idx}: Missing required fields.")
                skipped += 1
                continue

            if User.objects.filter(email__iexact=email).exists():
                errors.append(f"Row {idx}: Email '{email}' already exists.")
                skipped += 1
                continue
            if User.objects.filter(college_id__iexact=college_id).exists():
                errors.append(f"Row {idx}: College ID / Roll Number '{college_id}' already exists.")
                skipped += 1
                continue
            if User.objects.filter(roll_number__iexact=roll_number).exists():
                errors.append(f"Row {idx}: Roll number '{roll_number}' already exists.")
                skipped += 1
                continue

            from .models import School as SchoolModel, Program as ProgramModel, Branch as BranchModel
            from django.db.models import Q

            # Robust lookup matching either code OR full name case-insensitively
            school_obj = SchoolModel.objects.filter(
                Q(school_code__iexact=school_code) | Q(school_name__iexact=school_code)
            ).first()
            if not school_obj:
                errors.append(f"Row {idx}: School '{school_code}' not found.")
                skipped += 1
                continue

            program_obj = ProgramModel.objects.filter(
                school=school_obj
            ).filter(
                Q(program_code__iexact=program_code) | Q(program_name__iexact=program_code)
            ).first()
            if not program_obj:
                errors.append(f"Row {idx}: Program '{program_code}' not found in school '{school_code}'.")
                skipped += 1
                continue

            branch_obj = BranchModel.objects.filter(
                program=program_obj
            ).filter(
                Q(branch_code__iexact=branch_code) | Q(branch_name__iexact=branch_code)
            ).first()
            if not branch_obj:
                errors.append(f"Row {idx}: Branch '{branch_code}' not found in program '{program_code}'.")
                skipped += 1
                continue

            if year not in ["1", "2", "3", "4"]:
                year = "1"

            try:
                with db_transaction.atomic():
                    new_user = User.objects.create_user(
                        email=email,
                        password="itmu@123",
                        full_name=full_name,
                        roll_number=roll_number,
                        college_id=college_id,
                        role=User.Role.STUDENT,
                    )
                    StudentProfile.objects.create(
                        user=new_user,
                        school=school_obj,
                        program=program_obj,
                        branch=branch_obj,
                        year=year,
                    )
                created += 1
            except Exception as exc:
                errors.append(f"Row {idx}: {str(exc)}")
                skipped += 1

        return Response({
            "detail": f"Bulk upload complete. {created} created, {skipped} skipped.",
            "created": created,
            "skipped": skipped,
            "errors": errors[:50],
        })


class AdminStudentTemplateDownloadView(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request):
        user = get_user_from_request_token(request)
        if user is None or user.role != User.Role.ADMIN:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        from .models import School as SchoolModel, Program as ProgramModel, Branch as BranchModel
        from openpyxl.worksheet.datavalidation import DataValidation

        school_options = []
        for s in SchoolModel.objects.all():
            if s.school_code and s.school_code not in school_options:
                school_options.append(s.school_code)
            if s.school_name and s.school_name not in school_options:
                school_options.append(s.school_name)

        program_options = []
        for p in ProgramModel.objects.all():
            if p.program_code and p.program_code not in program_options:
                program_options.append(p.program_code)
            if p.program_name and p.program_name not in program_options:
                program_options.append(p.program_name)

        branch_options = []
        for b in BranchModel.objects.all():
            if b.branch_code and b.branch_code not in branch_options:
                branch_options.append(b.branch_code)
            if b.branch_name and b.branch_name not in branch_options:
                branch_options.append(b.branch_name)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        
        ws_lists = wb.create_sheet(title="AcademicSectors")
        ws_lists.sheet_state = 'hidden'
        
        ws_lists.append(["Schools", "Programs", "Branches"])
        
        max_len = max(len(school_options), len(program_options), len(branch_options))
        for i in range(max_len):
            ws_lists.append([
                school_options[i] if i < len(school_options) else "",
                program_options[i] if i < len(program_options) else "",
                branch_options[i] if i < len(branch_options) else ""
            ])
            
        headers = ["Full Name", "Roll Number", "Email", "School", "Program", "Branch", "Year"]
        ws.append(headers)
        ws.append(["John Doe", "ROLL001", "john@example.com", "School of Engineering & Technology", "BTECH", "CSE", "1"])

        school_range = f"$A$2:$A${len(school_options) + 1}" if school_options else "$A$2"
        program_range = f"$B$2:$B${len(program_options) + 1}" if program_options else "$B$2"
        branch_range = f"$C$2:$C${len(branch_options) + 1}" if branch_options else "$C$2"

        dv_school = DataValidation(type="list", formula1=f"=AcademicSectors!{school_range}", allow_blank=True)
        dv_program = DataValidation(type="list", formula1=f"=AcademicSectors!{program_range}", allow_blank=True)
        dv_branch = DataValidation(type="list", formula1=f"=AcademicSectors!{branch_range}", allow_blank=True)
        dv_year = DataValidation(type="list", formula1='"1,2,3,4"', allow_blank=True)

        ws.add_data_validation(dv_school)
        ws.add_data_validation(dv_program)
        ws.add_data_validation(dv_branch)
        ws.add_data_validation(dv_year)

        dv_school.add("D2:D1000")
        dv_program.add("E2:E1000")
        dv_branch.add("F2:F1000")
        dv_year.add("G2:G1000")

        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_length + 4, 15)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="student_upload_template.xlsx"'
        return response


class StudentChangePasswordView(APIView):
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        user = get_user_from_request_token(request)
        if user is None:
            return Response({"detail": "Authentication required."}, status=status.HTTP_401_UNAUTHORIZED)

        current_password = request.data.get("current_password", "").strip()
        new_password = request.data.get("new_password", "").strip()
        confirm_password = request.data.get("confirm_password", "").strip()

        if not current_password:
            return Response({"detail": "Current password is required."}, status=status.HTTP_400_BAD_REQUEST)
        if not user.check_password(current_password):
            return Response({"detail": "Current password is incorrect."}, status=status.HTTP_400_BAD_REQUEST)
        if not new_password or len(new_password) < 6:
            return Response({"detail": "New password must be at least 6 characters."}, status=status.HTTP_400_BAD_REQUEST)
        if new_password != confirm_password:
            return Response({"detail": "New password and confirmation do not match."}, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(new_password)
        user.rotate_session_token()
        user.save(update_fields=["password", "session_token", "updated_at"])

        return Response({
            "detail": "Password changed successfully. Please log in again.",
            "token": str(user.session_token),
        })


class SuperAdminManageAdminsView(APIView):
    authentication_classes = []
    permission_classes = []

    def _require_super_admin(self, request):
        user = get_user_from_request_token(request)
        if user is None or not user.is_super_admin:
            return None
        return user

    def get(self, request):
        user = get_user_from_request_token(request)
        if user is None or user.role != User.Role.ADMIN:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
            
        if user.is_super_admin:
            admins = User.objects.filter(role=User.Role.ADMIN).select_related('school')
        else:
            if not user.school:
                admins = User.objects.filter(id=user.id)
            else:
                admins = User.objects.filter(role=User.Role.ADMIN, school=user.school).select_related('school')
                
        data = []
        for a in admins:
            data.append({
                "id": a.id,
                "full_name": a.full_name,
                "email": a.email,
                "college_id": a.college_id,
                "school_id": a.school.id if a.school else None,
                "school_name": a.school.school_name if a.school else "Global / Super Admin" if a.is_super_admin else "None",
                "school_code": a.school.school_code if a.school else "SUPER" if a.is_super_admin else "",
                "is_super_admin": a.is_super_admin,
                "is_active": a.is_active,
                "cleartext_password": a.cleartext_password or "",
            })
        return Response(data)

    def post(self, request):
        if not self._require_super_admin(request):
            return Response({"detail": "Super Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        full_name = request.data.get("full_name", "").strip()
        email = request.data.get("email", "").strip().lower()
        college_id = request.data.get("college_id", "").strip().upper()
        school_id = request.data.get("school_id")
        password = request.data.get("password", "").strip()
        is_super = request.data.get("is_super_admin", False)

        if not full_name or not email or not college_id or not password:
            return Response({"detail": "full_name, email, college_id, and password are required."}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(email__iexact=email).exists():
            return Response({"detail": "A user with this email already exists."}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(college_id__iexact=college_id).exists():
            return Response({"detail": "A user with this college ID already exists."}, status=status.HTTP_400_BAD_REQUEST)

        school_obj = None
        if school_id:
            try:
                school_obj = School.objects.get(id=school_id)
            except School.DoesNotExist:
                return Response({"detail": "Selected school not found."}, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.create_user(
            email=email,
            password=password,
            full_name=full_name,
            college_id=college_id,
            roll_number=college_id,
            role=User.Role.ADMIN,
            school=school_obj,
            is_super_admin=is_super,
            cleartext_password=password,
            is_staff=True,
            is_superuser=is_super
        )

        return Response({"detail": "Admin user created successfully.", "id": user.id}, status=status.HTTP_201_CREATED)


class SuperAdminManageAdminsDetailView(APIView):
    authentication_classes = []
    permission_classes = []

    def _require_super_admin(self, request):
        user = get_user_from_request_token(request)
        if user is None or not user.is_super_admin:
            return None
        return user

    def put(self, request, pk):
        if not self._require_super_admin(request):
            return Response({"detail": "Super Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        try:
            admin = User.objects.get(pk=pk, role=User.Role.ADMIN)
        except User.DoesNotExist:
            return Response({"detail": "Admin user not found."}, status=status.HTTP_404_NOT_FOUND)

        full_name = request.data.get("full_name", "").strip()
        email = request.data.get("email", "").strip().lower()
        college_id = request.data.get("college_id", "").strip().upper()
        school_id = request.data.get("school_id")
        password = request.data.get("password", "").strip()
        is_super = request.data.get("is_super_admin", admin.is_super_admin)
        is_active = request.data.get("is_active", admin.is_active)

        if full_name:
            admin.full_name = full_name
        if email:
            if User.objects.exclude(id=admin.id).filter(email__iexact=email).exists():
                return Response({"detail": "A user with this email already exists."}, status=status.HTTP_400_BAD_REQUEST)
            admin.email = email
        if college_id:
            if User.objects.exclude(id=admin.id).filter(college_id__iexact=college_id).exists():
                return Response({"detail": "A user with this college ID already exists."}, status=status.HTTP_400_BAD_REQUEST)
            admin.college_id = college_id
            admin.roll_number = college_id

        if 'school_id' in request.data:
            if school_id:
                try:
                    admin.school = School.objects.get(id=school_id)
                except School.DoesNotExist:
                    return Response({"detail": "Selected school not found."}, status=status.HTTP_400_BAD_REQUEST)
            else:
                admin.school = None

        if password:
            admin.set_password(password)
            admin.cleartext_password = password

        admin.is_super_admin = is_super
        admin.is_superuser = is_super
        admin.is_active = is_active
        admin.save()

        return Response({"detail": "Admin user updated successfully."})

    def delete(self, request, pk):
        if not self._require_super_admin(request):
            return Response({"detail": "Super Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        try:
            admin = User.objects.get(pk=pk, role=User.Role.ADMIN)
        except User.DoesNotExist:
            return Response({"detail": "Admin user not found."}, status=status.HTTP_404_NOT_FOUND)

        if admin == get_user_from_request_token(request):
            return Response({"detail": "You cannot delete your own account."}, status=status.HTTP_400_BAD_REQUEST)

        admin.delete()
        return Response({"detail": "Admin user deleted successfully."})
