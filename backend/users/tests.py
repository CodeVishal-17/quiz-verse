from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework import status
from django.contrib.auth import get_user_model
from .models import School, Program, Branch, StudentProfile
import io
import openpyxl

User = get_user_model()

class UserAuthAndAdminTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.school = School.objects.create(school_name="School of Engineering", school_code="SOET")
        self.program = Program.objects.create(school=self.school, program_name="Computer Science", program_code="CSE_PROG")
        self.branch = Branch.objects.create(program=self.program, branch_name="Computer Science Engineering", branch_code="CSE")

        # Create standard student
        self.student = User.objects.create_user(
            email="aria@example.edu",
            password="strongpass123",
            full_name="Aria Sharma",
            college_id="ROLL001",
            roll_number="ROLL001",
            role=User.Role.STUDENT
        )
        StudentProfile.objects.create(
            user=self.student,
            school=self.school,
            program=self.program,
            branch=self.branch,
            year="3"
        )

        # Create admin user
        self.admin = User.objects.create_superuser(
            email="admin@example.edu",
            password="adminpassword123",
            full_name="Admin User",
            college_id="ADMIN001",
            roll_number="ADMIN001"
        )

    def test_student_login_success(self):
        # Student login using roll_number / email
        payload = {
            "identifier": "ROLL001",
            "password": "strongpass123"
        }
        response = self.client.post("/api/users/login/", payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("token", response.data)
        self.assertEqual(response.data["student"]["email"], self.student.email)

    def test_student_login_failure(self):
        payload = {
            "identifier": "ROLL001",
            "password": "wrongpassword"
        }
        response = self.client.post("/api/users/login/", payload, format="json")
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_download_template_with_dropdowns(self):
        url = "/api/users/admin/students/download-template/"
        response = self.client.get(url, HTTP_AUTHORIZATION=f"Bearer {self.admin.session_token}")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Load sheet to verify content
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertIn("Students", wb.sheetnames)
        self.assertIn("AcademicSectors", wb.sheetnames)

        ws = wb["Students"]
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, ["Full Name", "Roll Number", "Email", "School", "Program", "Branch", "Year"])

        # Check hidden data lists
        ws_lists = wb["AcademicSectors"]
        list_headers = [cell.value for cell in ws_lists[1]]
        self.assertEqual(list_headers, ["Schools", "Programs", "Branches"])

        # Verify our created school, program, and branch are in the hidden sheet
        schools_col = [cell.value for cell in ws_lists["A"]]
        self.assertIn("SOET", schools_col)
        self.assertIn("School of Engineering", schools_col)

        programs_col = [cell.value for cell in ws_lists["B"]]
        self.assertIn("CSE_PROG", programs_col)

        branches_col = [cell.value for cell in ws_lists["C"]]
        self.assertIn("CSE", branches_col)
